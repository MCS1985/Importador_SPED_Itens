import openpyxl
from collections import Counter

def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    raw_headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {}
    for i, h in enumerate(raw_headers):
        header_map[h.lower().replace('\u00a0', ' ')] = i
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for nh, idx in header_map.items():
            if idx < len(row):
                d[nh] = row[idx]
        rows.append(d)
    return header_map, rows

def safe_str(v):
    if v is None: return ''
    s = str(v).strip()
    if s.lower() in ('none', 'nan', ''): return ''
    return s

path1 = r'D:\FERRAMENTAS_MCS_IA\Importador_SPED_Com_Itens\Relatorio_CFOP_2026-06-21-21-25-15.xlsx'
path2 = r'D:\FERRAMENTAS_MCS_IA\Importador_SPED_Com_Itens\Relatorio_CFOP_SAAM - FISCAL - Saídas - C170.xlsx'

hm1, r1 = read_xlsx(path1)
hm2, r2 = read_xlsx(path2)

# Tool: period distribution
print('=== PERIODOS NO TOOL ===')
tool_periodos = Counter()
for r in r1:
    p = safe_str(r.get('per\u00edodo', ''))
    if p: tool_periodos[p] += 1
for p, c in sorted(tool_periodos.items()):
    print('  {}: {} itens'.format(p, c))

print()
print('Total tool: {}'.format(len(r1)))

# SAAM: mes distribution  
print()
print('=== MESES NO SAAM ===')
saam_meses = Counter()
for r in r2:
    m = safe_str(r.get('m\u00eas de refer\u00eancia', safe_str(r.get('m\u00eas de refer\u00eancia', ''))))
    if m: saam_meses[m] += 1
    else:
        # Try to extract from chave
        ch = safe_str(r.get('chave', ''))
        if len(ch) >= 6:
            mes = ch[4:6] + '/' + ch[2:4]
            saam_meses[mes] += 1
        else:
            saam_meses['SEM DATA'] += 1
for m, c in sorted(saam_meses.items()):
    print('  {}: {} itens'.format(m, c))
print()
print('Total SAAM: {}'.format(len(r2)))

# Also check what data exists in tool for CFOP 5102
print()
print('=== CFOPs NO TOOL ===')
tool_cfops = Counter()
for r in r1:
    cfop = safe_str(r.get('cfop', ''))
    if cfop: tool_cfops[cfop] += 1
for cf, c in tool_cfops.most_common(20):
    print('  CFOP {}: {} itens'.format(cf, c))

print()
print('=== CFOPs NO SAAM (top 20) ===')
saam_cfops = Counter()
for r in r2:
    cfop = safe_str(r.get('cfop', ''))
    if cfop: saam_cfops[cfop] += 1
for cf, c in saam_cfops.most_common(20):
    print('  CFOP {}: {} itens'.format(cf, c))

# Check tool data for 02/2025
print()
print('=== ITENS DO TOOL COM PERIODO 02/2025 ===')
count = 0
for r in r1:
    p = safe_str(r.get('per\u00edodo', ''))
    if '02/2025' in p:
        ch = safe_str(r.get('chave de acesso', ''))
        ni = safe_str(r.get('n\u00famero do item', ''))
        cf = safe_str(r.get('cfop', ''))
        vl = r.get('valor do item', 0) or 0
        print('  chave={}, nItem={}, cfop={}, vl={}'.format(ch, ni, cf, vl))
        count += 1
if count == 0:
    print('  NENHUM ITEM DE 02/2025 NO TOOL!')
