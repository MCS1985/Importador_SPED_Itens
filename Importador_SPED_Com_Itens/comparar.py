import openpyxl

# Read SAAM
f1 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_SAAM.xlsx'
wb1 = openpyxl.load_workbook(f1, read_only=True, data_only=True)
ws1 = wb1.active
rows1 = list(ws1.iter_rows(min_row=2, values_only=True))
saam = {}
for r in rows1:
    chave = str(r[5] or '').strip()
    num_item = str(r[19] or '').strip()
    key = chave + '|' + num_item
    if key != '|':
        saam[key] = {
            'chave': chave,
            'num_item': num_item,
            'vl_item': r[27] or 0,
            'cst_pis': str(r[47] or '').strip(),
            'cfop': str(r[31] or '').strip(),
        }

# Read Personalizado
f2 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_2026-06-21-15-30-43.xlsx'
wb2 = openpyxl.load_workbook(f2, read_only=True, data_only=True)
ws2 = wb2.active
rows2 = list(ws2.iter_rows(min_row=5, values_only=True))
personal = {}
for r in rows2:
    if not r[1] or not r[5]: continue
    if r[1] == 'Chave unica': continue
    chave = str(r[5] or '').strip()
    num_item = str(r[9] or '').strip()
    key = chave + '|' + num_item
    personal[key] = {
        'chave': chave,
        'num_item': num_item,
        'vl_item': r[14] or 0,
    }

print('SAAM itens:', len(saam))
print('Personalizado itens:', len(personal))

# Find missing in Personalizado
missing = []
for key, item in saam.items():
    if key not in personal:
        missing.append(item)

print('\nItens no SAAM que NAO estao no Personalizado:', len(missing))
total = sum(float(m['vl_item'] or 0) for m in missing)
print('Valor total faltante: R$ {:.2f}'.format(total))

# Group by CFOP
cfop_groups = {}
for m in missing:
    cfop = m['cfop']
    if cfop not in cfop_groups:
        cfop_groups[cfop] = {'count': 0, 'total': 0}
    cfop_groups[cfop]['count'] += 1
    cfop_groups[cfop]['total'] += float(m['vl_item'] or 0)

print('\n--- Por CFOP ---')
for cfop, g in sorted(cfop_groups.items(), key=lambda x: -x[1]['total']):
    print('  CFOP {}: {} itens, R$ {:.2f}'.format(cfop, g['count'], g['total']))

# Group by CST PIS
cst_groups = {}
for m in missing:
    cst = m['cst_pis']
    if cst not in cst_groups:
        cst_groups[cst] = {'count': 0, 'total': 0}
    cst_groups[cst]['count'] += 1
    cst_groups[cst]['total'] += float(m['vl_item'] or 0)

print('\n--- Por CST PIS ---')
for cst, g in sorted(cst_groups.items(), key=lambda x: -x[1]['total']):
    print('  CST {}: {} itens, R$ {:.2f}'.format(cst, g['count'], g['total']))

# Show some sample missing items
print('\n--- Primeiros 15 itens faltantes ---')
for m in missing[:15]:
    ch = m['chave']
    if len(ch) > 25: ch = ch[:25] + '...'
    print('  Chave={}\tItem={}\tCFOP={}\tCST={}\tR$ {:.2f}'.format(
        ch, m['num_item'], m['cfop'], m['cst_pis'], float(m['vl_item'] or 0)))
if len(missing) > 15:
    print('  ... e mais', len(missing) - 15, 'itens')

# Also check: items in Personalizado but not in SAAM
extra = []
for key, item in personal.items():
    if key not in saam:
        extra.append(item)
print('\nItens no Personalizado que NAO estao no SAAM:', len(extra))
if extra:
    total_extra = sum(float(m['vl_item'] or 0) for m in extra)
    print('Valor total extra: R$ {:.2f}'.format(total_extra))
