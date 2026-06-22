import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook('doc1782060517220.xlsx', data_only=True)
ws = wb.active
print(f'Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}')
print()
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
    vals = [str(v)[:50] if v is not None else '' for v in row]
    print(f'{i}: ' + ' | '.join(vals))
