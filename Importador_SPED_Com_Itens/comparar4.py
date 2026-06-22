import openpyxl
from collections import defaultdict

# ---- SAAM XLSX ----
f1 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_SAAM.xlsx'
wb1 = openpyxl.load_workbook(f1, read_only=True, data_only=True)
ws1 = wb1.active
headers1 = [str(c.value) for c in next(ws1.iter_rows(min_row=1, max_row=1))]
rows1 = list(ws1.iter_rows(min_row=2, values_only=True))

saam_por_grupo = defaultdict(lambda: {'count': 0, 'total': 0, 'itens': []})
saam_keys_set = set()
for r in rows1:
    chave = str(r[5] or '').strip()
    num_item = str(r[19] or '').strip()
    key = chave + '|' + num_item
    if key == '|':
        continue
    saam_keys_set.add(key)
    vl = float(r[27] or 0)
    sentido = str(r[6] or '').strip()
    cfop = str(r[31] or '').strip()
    cst = str(r[47] or '').strip()
    grupo = (sentido, cfop[:4])
    saam_por_grupo[grupo]['count'] += 1
    saam_por_grupo[grupo]['total'] += vl
    saam_por_grupo[grupo]['itens'].append((key, cfop, cst, vl))

# ---- PERSONALIZADO XLSX ----
f2 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_2026-06-21-15-30-43.xlsx'
wb2 = openpyxl.load_workbook(f2, read_only=True, data_only=True)
ws2 = wb2.active

header_row = None
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=10, values_only=True)):
    if row[1] == 'Chave unica':
        header_row = i + 1
        break

rows2 = list(ws2.iter_rows(min_row=header_row + 1, values_only=True))

pers_por_grupo = defaultdict(lambda: {'count': 0, 'total': 0, 'itens': []})
pers_keys_set = set()
for r in rows2:
    if not r[1] or not r[5]:
        continue
    if r[1] == 'Chave unica':
        continue
    chave = str(r[5] or '').strip()
    num_item = str(r[9] or '').strip()
    key = chave + '|' + num_item
    if key == '|':
        continue
    pers_keys_set.add(key)
    vl = float(r[14] or 0)
    sentido = str(r[6] or '').strip()
    registro = str(r[7] or '').strip()
    cfop = str(r[12] or '').strip()
    grupo = (sentido, cfop[:4])
    pers_por_grupo[grupo]['count'] += 1
    pers_por_grupo[grupo]['total'] += vl
    pers_por_grupo[grupo]['itens'].append((key, cfop, registro, vl))

# ---- Key statistics ----
print('=== ESTATISTICAS DE CHAVE ===')
print(f'SAAM keys unicas:        {len(saam_keys_set)}')
print(f'Personalizado keys unicas: {len(pers_keys_set)}')
print()

# ---- Compare by group (same group key now) ----
print('=== COMPARACAO POR (SENTIDO, CFOP) ===')
print(f'{"Sentido":10s} {"CFOP":6s} {"SAAM #":>6s} {"Pers #":>6s} {"Diff #":>6s} {"SAAM R$":>12s} {"Pers R$":>12s} {"Diff R$":>12s}')
print('-' * 80)

all_grupos = set(list(saam_por_grupo.keys()) + list(pers_por_grupo.keys()))
for grupo in sorted(all_grupos):
    s = saam_por_grupo.get(grupo, {'count': 0, 'total': 0})
    p = pers_por_grupo.get(grupo, {'count': 0, 'total': 0})
    if s['count'] == 0 and p['count'] == 0:
        continue
    print(f'{grupo[0]:10s} {grupo[1]:6s} {s["count"]:6d} {p["count"]:6d} {p["count"] - s["count"]:6d}  R${s["total"]:>10.2f}  R${p["total"]:>10.2f}  R${p["total"] - s["total"]:>10.2f}')

saam_total = sum(g['count'] for g in saam_por_grupo.values())
pers_total = sum(g['count'] for g in pers_por_grupo.values())
saam_vl = sum(g['total'] for g in saam_por_grupo.values())
pers_vl = sum(g['total'] for g in pers_por_grupo.values())
print(f'{"TOTAL":10s} {"":6s} {saam_total:6d} {pers_total:6d} {pers_total - saam_total:6d}  R${saam_vl:>10.2f}  R${pers_vl:>10.2f}  R${pers_vl - saam_vl:>10.2f}')
print()

# ---- Find items in SAAM not in personalizado by KEY ----
print('--- Itens no SAAM que NAO estao no Personalizado (por chave+num_item) ---')
missing = saam_keys_set - pers_keys_set
print(f'Total: {len(missing)} itens')
for mk in sorted(list(missing)[:30]):
    for g, data in saam_por_grupo.items():
        for k, cfop, cst, vl in data['itens']:
            if k == mk:
                print(f'  CFOP={cfop:6s} CST={cst:4s} Item={k[-5:]:5s} R${vl:>10.2f}')
                break
if len(missing) > 30:
    print(f'  ... e mais {len(missing) - 30}')

extra = pers_keys_set - saam_keys_set
print(f'\n--- Itens no Personalizado que NAO estao no SAAM: {len(extra)} ---')
for ek in sorted(list(extra)[:30]):
    for g, data in pers_por_grupo.items():
        for k, cfop, reg, vl in data['itens']:
            if k == ek:
                print(f'  {reg:5s} CFOP={cfop:6s} Item={k[-5:]:5s} R${vl:>10.2f}')
                break
if len(extra) > 30:
    print(f'  ... e mais {len(extra) - 30}')
print()

# ---- Cross-reference: items with SAME key but DIFFERENT CFOP ----
print('=== ITENS COM MESMA CHAVE MAS CFOP DIFERENTE ===')
saam_by_key = {}
for g, data in saam_por_grupo.items():
    for k, cfop, cst, vl in data['itens']:
        saam_by_key[k] = {'cfop': cfop, 'cst': cst, 'vl': vl, 'grupo': g}

pers_by_key = {}
for g, data in pers_por_grupo.items():
    for k, cfop, reg, vl in data['itens']:
        pers_by_key[k] = {'cfop': cfop, 'registro': reg, 'vl': vl, 'grupo': g}

common_keys = saam_keys_set & pers_keys_set
mismatch_count = 0
for k in sorted(common_keys):
    saam_info = saam_by_key[k]
    pers_info = pers_by_key[k]
    if saam_info['cfop'][:4] != pers_info['cfop'][:4]:
        mismatch_count += 1
        if mismatch_count <= 30:
            print(f'  Chave={k[:25]:25s} Item={k[-5:]:5s} SAAM CFOP={saam_info["cfop"]:6s} Pers CFOP={pers_info["cfop"]:6s} SAAM R${saam_info["vl"]:>10.2f} Pers R${pers_info["vl"]:>10.2f}')

print(f'Total de itens com CFOP diferente: {mismatch_count}')
