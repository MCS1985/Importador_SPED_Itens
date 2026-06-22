import openpyxl
from collections import Counter

# Read SAAM
f1 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_SAAM.xlsx'
wb1 = openpyxl.load_workbook(f1, read_only=True, data_only=True)
ws1 = wb1.active
headers = [str(c.value) for c in next(ws1.iter_rows(min_row=1, max_row=1))]
rows1 = list(ws1.iter_rows(min_row=2, values_only=True))

print('=== SAAM ===')
print('Headers:', headers)
print('Rows:', len(rows1))

# Distribution by CFOP (col 32), CST (col 48)
cfop_count = Counter()
cst_count = Counter()
sem_chave = 0
total_vl = 0
for r in rows1:
    chave = str(r[5] or '').strip()
    cfop = str(r[31] or '').strip()
    cst = str(r[47] or '').strip()
    vl = float(r[27] or 0)
    if not chave: sem_chave += 1
    elif cfop: cfop_count[cfop] += 1
    if cst: cst_count[cst] += 1
    total_vl += vl

print('Items sem chave:', sem_chave)
print('Valor total: R$ {:.2f}'.format(total_vl))
print()

print('--- Top 20 CFOP ---')
for cfop, n in cfop_count.most_common(20):
    print('  CFOP {}: {}'.format(cfop, n))

print()
print('--- Top 20 CST PIS ---')
for cst, n in cst_count.most_common(20):
    print('  CST {}: {}'.format(cst, n))

# Also check: does the personalizado have C170 CONTRIB + C170 FISCAL?
f2 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_2026-06-21-15-30-43.xlsx'
wb2 = openpyxl.load_workbook(f2, read_only=True, data_only=True)
ws2 = wb2.active
rows2 = list(ws2.iter_rows(min_row=5, values_only=True))

print('\n=== PERSONALIZADO ===')
reg_count = Counter()
sit_count = Counter()
cfop_p = Counter()
total_vl2 = 0
for r in rows2:
    if not r[1] or not r[5]: continue
    if r[1] == 'Chave unica': continue
    reg = str(r[7] or '').strip()
    sit = str(r[6] or '').strip()
    cfop = str(r[11] or '').strip()
    vl = float(r[14] or 0)
    if reg: reg_count[reg] += 1
    if sit: sit_count[sit] += 1
    if cfop: cfop_p[cfop] += 1
    total_vl2 += vl

print('Registros:', dict(reg_count))
print('Sentido:', dict(sit_count))
print('Valor total: R$ {:.2f}'.format(total_vl2))

print()
print('--- Top 20 CFOP ---')
for cfop, n in cfop_p.most_common(20):
    print('  CFOP {}: {}'.format(cfop, n))
