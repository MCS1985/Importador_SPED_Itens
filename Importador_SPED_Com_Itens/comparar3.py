import openpyxl
from collections import defaultdict

# ---- SAAM XLSX ----
f1 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_SAAM.xlsx'
wb1 = openpyxl.load_workbook(f1, read_only=True, data_only=True)
ws1 = wb1.active
headers1 = [str(c.value) for c in next(ws1.iter_rows(min_row=1, max_row=1))]
rows1 = list(ws1.iter_rows(min_row=2, values_only=True))

# SAAM columns: 5=Chave, 6=IndOper (Entrada/Saida), 19=NumItem, 27=ValorItem, 31=CFOP, 47=CST_PIS
saam_por_grupo = defaultdict(lambda: {'count': 0, 'total': 0, 'itens': []})
for r in rows1:
    chave = str(r[5] or '').strip()
    num_item = str(r[19] or '').strip()
    key = chave + '|' + num_item
    if key == '|':
        continue
    vl = float(r[27] or 0)
    sentido = str(r[6] or '').strip()
    cfop = str(r[31] or '').strip()
    cst = str(r[47] or '').strip()
    # SAAM nao tem coluna de registro ou sped_origem, entao agrupamos por sentido+cfop
    grupo = ('SAAM', sentido, cfop[:4])
    saam_por_grupo[grupo]['count'] += 1
    saam_por_grupo[grupo]['total'] += vl
    saam_por_grupo[grupo]['itens'].append(key)

# ---- PERSONALIZADO XLSX ----
f2 = 'D:\\FERRAMENTAS_MCS_IA\\Importador_SPED_Com_Itens\\Relatorio_CFOP_2026-06-21-15-30-43.xlsx'
wb2 = openpyxl.load_workbook(f2, read_only=True, data_only=True)
ws2 = wb2.active

# Find header row (it has 'Chave unica')
header_row = None
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=10, values_only=True)):
    if row[1] == 'Chave unica':
        header_row = i + 1
        break

rows2 = list(ws2.iter_rows(min_row=header_row + 1, values_only=True))

# Personalizado columns: 1=ChaveUnica, 2=Periodo, 3=Data, 4=CNPJ, 5=ChaveAcesso, 6=Sentido, 7=Registro, 9=NumItem, 11=CFOP (description), 12=CFOP? 
# Let me find the correct columns
print('Personalizado header row:', header_row)
# Re-read header row
for i, row in enumerate(ws2.iter_rows(min_row=header_row, max_row=header_row, values_only=True)):
    for j, v in enumerate(row):
        if v:
            print(f'  Col {j}: {v}')
print()

pers_por_grupo = defaultdict(lambda: {'count': 0, 'total': 0, 'itens': []})
for r in rows2:
    if not r[1] or not r[5]:
        continue
    if r[1] == 'Chave unica':
        continue
    chave = str(r[5] or '').strip()
    num_item = str(r[9] or '').strip()
    key = chave + '|' + num_item
    vl = float(r[14] or 0)
    sentido = str(r[6] or '').strip()
    registro = str(r[7] or '').strip()
    # CFOP is at column index 12 (based on COLUMNS definition: cfop is after descricao at index 11)
    # Actually looking at COLUMNS: chave_unica(0), periodo(1), dt_e_s(2), cnpj_empresa(3), chv_doce(4), 
    # sentido(5), registro(6), num_doc(7), num_item(8), cod_item(9), descricao(10), cfop(11), ncm(12),
    # vl_item(13), vl_desc(14)...
    cfop = str(r[12] or '').strip()
    grupo = (registro, sentido, cfop[:4])
    pers_por_grupo[grupo]['count'] += 1
    pers_por_grupo[grupo]['total'] += vl
    pers_por_grupo[grupo]['itens'].append(key)

# Compare
print('=== COMPARACAO ===')
print()

# Total counts
saam_total = sum(g['count'] for g in saam_por_grupo.values())
pers_total = sum(g['count'] for g in pers_por_grupo.values())
saam_vl = sum(g['total'] for g in saam_por_grupo.values())
pers_vl = sum(g['total'] for g in pers_por_grupo.values())
print(f'SAAM:        {saam_total} itens, R$ {saam_vl:.2f}')
print(f'Personalizado: {pers_total} itens, R$ {pers_vl:.2f}')
print(f'Diferenca:   {pers_total - saam_total} itens, R$ {pers_vl - saam_vl:.2f}')
print()

# Compare by (sentido, cfop prefix) groups
print('--- Por (Sentido, CFOP) ---')
print(f'{"Sentido":10s} {"CFOP":6s} {"SAAM #":>6s} {"Pers #":>6s} {"Diff #":>6s} {"SAAM R$":>12s} {"Pers R$":>12s} {"Diff R$":>12s}')
print('-' * 80)

all_grupos = set(list(saam_por_grupo.keys()) + list(pers_por_grupo.keys()))
for grupo in sorted(all_grupos):
    s = saam_por_grupo.get(grupo, {'count': 0, 'total': 0})
    p = pers_por_grupo.get(grupo, {'count': 0, 'total': 0})
    if s['count'] == 0 and p['count'] == 0:
        continue
    print(f'{grupo[1]:10s} {grupo[2]:6s} {s["count"]:6d} {p["count"]:6d} {p["count"] - s["count"]:6d}  R${s["total"]:>10.2f}  R${p["total"]:>10.2f}  R${p["total"] - s["total"]:>10.2f}')

# Find items in SAAM not in personalizado
print('\n--- Itens no SAAM que NAO estao no Personalizado ---')
saam_keys = set()
for g, data in saam_por_grupo.items():
    for k in data['itens']:
        saam_keys.add(k)

pers_keys = set()
for g, data in pers_por_grupo.items():
    for k in data['itens']:
        pers_keys.add(k)

missing = saam_keys - pers_keys
print(f'Total: {len(missing)} itens')
for mk in sorted(list(missing)[:30]):
    # Find in SAAM
    for r in rows1:
        ch = str(r[5] or '').strip()
        ni = str(r[19] or '').strip()
        if ch + '|' + ni == mk:
            print(f'  Chave={ch[:25]:25s} Item={ni:5s} CFOP={str(r[31] or ""):6s} CST={str(r[47] or ""):4s} R${float(r[27] or 0):>10.2f}')
            break
if len(missing) > 30:
    print(f'  ... e mais {len(missing) - 30}')

extra = pers_keys - saam_keys
print(f'\nItens no Personalizado que NAO estao no SAAM: {len(extra)}')
